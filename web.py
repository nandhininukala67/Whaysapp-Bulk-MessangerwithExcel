import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from urllib.parse import quote
import os

options = Options()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_argument("--profile-directory=Default")
options.add_argument("--user-data-dir=/var/tmp/chrome_user_data")

os.system("")
os.environ["WDM_LOG_LEVEL"] = "0"
class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

print(style.BLUE)
print("**********************************************************")
print("**********************************************************")
print("*****                                               ******")
print("*****  Initiating Mass Messenger >WhatsApp<         ******")
print("*****                                               ******")
print("**********************************************************")
print("**********************************************************")
print(style.RESET)

f = open("message.txt", "r")
message = f.read()
f.close()

print(style.YELLOW + '\nThis is your message-')
print(style.GREEN + message)
print("\n" + style.RESET)
message = quote(message)

numbers = []
b = openpyxl.load_workbook("numbers.xlsx")

sht = b.active

for col in sht.iter_cols(1, sht.max_column):
    for i in range(1, sht.max_row):
        numbers.append(col[i].value)
    break
f.close()
total_number = len(numbers)

name = []
b = openpyxl.load_workbook("numbers.xlsx")

sht = b.active

for col in sht.iter_cols(2, sht.max_column):
    for i in range(1, sht.max_row):
        name.append(col[i].value)
    break
f.close()
total_number = len(name)
print(style.RED + 'We found ' + str(total_number) + ' numbers in the file' + style.RESET)
delay = 30

driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
print('Once your browser opens up sign in to web whatsapp')
driver.get('https://web.whatsapp.com')
input(style.MAGENTA + "AFTER logging into Whatsapp Web is complete and your chats are visible, press ENTER..." + style.RESET)
for idx, number in enumerate(numbers):
	#number = number.strip()
	if number == "":
		continue
	print(style.YELLOW + '{}/{} => Sending message to {}.'.format((idx+1), total_number, number) + style.RESET)
	try:
		url = 'https://web.whatsapp.com/send?phone=' + "91"+str(number) + '&text=' " "'Hello'+  "  "  +str(name[idx])+ "  "+os.linesep+'\n'+message
		sent = False
		for i in range(3):
			if not sent:
				driver.get(url)
				try:
					click_btn = WebDriverWait(driver, delay).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='compose-btn-send']")))
				except Exception as e:
					print(style.RED + f"\nFailed to send message to: {number}, retry ({i+1}/3)")
					print("Make sure your phone and computer is connected to the internet.")
					print("If there is an alert, please dismiss it." + style.RESET)
				else:
					sleep(1)
					click_btn.click()
					sent=True
					sleep(3)
					print(style.GREEN + 'Message sent to: ' + str(number) +"  "+str(name[idx])+ style.RESET)
	except Exception as e:
		print(style.RED + 'Failed to send message to ' + str(number) + ""+str(name[idx])+str(e) + style.RESET)
driver.close()